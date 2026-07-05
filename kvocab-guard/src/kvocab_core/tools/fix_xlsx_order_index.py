"""Fix order_index in xlsx to new formula (201xxx for 2A)."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from kvocab_core.config import compute_order_index

ROOT = Path(__file__).resolve().parents[3]
PATHS = [
    ROOT / "data" / "seed" / "vocabulary_database.xlsx",
    ROOT.parent / "snu2a_level_mapped_vocabulary.xlsx",
    ROOT.parent / "snu2b_level_mapped_vocabulary.xlsx",
    ROOT.parent / "snu3a_level_mapped_vocabulary.xlsx",
]


def fix_path(path: Path) -> None:
    if not path.exists():
        return
    wb = openpyxl.load_workbook(path)
    for sheet_name in ("Level_Map", "Vocabulary"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {name: i + 1 for i, name in enumerate(header)}
        for row in ws.iter_rows(min_row=2):
            level = row[col["level"] - 1].value
            unit_no = row[col["unit_no"] - 1].value
            lesson_no = row[col["lesson_no"] - 1].value
            if level and unit_no is not None and lesson_no is not None:
                oi = compute_order_index(str(level), int(unit_no), int(lesson_no))
                row[col["order_index"] - 1].value = oi
                if "lesson" in col:
                    row[col["lesson"] - 1].value = f"{int(unit_no)}-{int(lesson_no)}"
    wb.save(path)
    print(f"Updated {path}")


if __name__ == "__main__":
    for p in PATHS:
        fix_path(p)
