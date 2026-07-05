from __future__ import annotations

import json
import logging
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from kvocab_core.config import compute_order_index
from kvocab_core.models import Lesson, Lexeme, Occurrence, SurfaceForm, UnmappedStaging
from kvocab_core.normalization import normalize_key
from kvocab_core.token_units import build_match_segments, segment_key

logger = logging.getLogger(__name__)

VOCAB_HEADERS = [
    "level",
    "unit_no",
    "lesson_no",
    "unit_topic",
    "unit_title",
    "first_page",
    "order_index",
    "lemma",
    "gloss_en",
    "source_type",
]


def lesson_code(unit_no: int | str, lesson_no: int | str) -> str:
    return f"{int(unit_no)}-{int(lesson_no)}"


def _lesson_lookup(session: Session, level: str) -> dict[str, Lesson]:
    rows = session.query(Lesson).filter(Lesson.level == level).all()
    return {r.lesson: r for r in rows}


def _page_lesson_map(session: Session, level: str) -> list[Lesson]:
    return session.query(Lesson).filter(Lesson.level == level).order_by(Lesson.order_index).all()


def lesson_for_page(lessons: list[Lesson], page: int) -> Lesson | None:
    for ls in lessons:
        if ls.page_start <= page <= ls.page_end:
            return ls
    return None


def _stage_unmapped(
    session: Session,
    *,
    lemma: str | None,
    level: str | None,
    lesson: str | None,
    first_page: int | None,
    reason: str,
    metadata: dict | None = None,
) -> None:
    session.add(
        UnmappedStaging(
            lemma=lemma,
            level=level,
            lesson=lesson,
            first_page=first_page,
            reason=reason,
            raw_source_metadata=json.dumps(metadata or {}, ensure_ascii=False),
        )
    )


def _metadata_from_row(row: dict) -> dict:
    return {
        k: row.get(k)
        for k in ("unit_topic", "unit_title")
        if row.get(k) not in (None, "")
    }


def import_level_map_sheet(session: Session, ws) -> int:
    count = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        level, unit_no, lesson_no = r[0], r[1], r[2]
        lesson = lesson_code(unit_no, lesson_no)
        _page_start, _page_end = r[5], r[6]
        stored_oi = r[7] if len(r) > 7 else None
        calc_oi = compute_order_index(str(level), int(unit_no), int(lesson_no))
        if stored_oi is not None and int(stored_oi) != calc_oi:
            logger.warning(
                "Level_Map order_index mismatch %s %s: file=%s calc=%s",
                level,
                lesson,
                stored_oi,
                calc_oi,
            )
        ls = (
            session.query(Lesson)
            .filter(Lesson.level == level, Lesson.lesson == lesson)
            .one_or_none()
        )
        if ls:
            ls.order_index = calc_oi
            ls.page_start = int(_page_start)
            ls.page_end = int(_page_end)
        count += 1
    session.commit()
    return count


def import_vocabulary_xlsx(session: Session, path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Level_Map" in wb.sheetnames:
        import_level_map_sheet(session, wb["Level_Map"])

    ws = wb["Vocabulary"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: i for i, name in enumerate(header)}

    stats = {
        "imported_lexemes": 0,
        "occurrences": 0,
        "unmapped": 0,
        "warnings": [],
    }

    lessons_by_level: dict[str, list[Lesson]] = {}

    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not row_vals or not any(row_vals):
            continue

        row = {h: (row_vals[col[h]] if col.get(h) is not None else None) for h in VOCAB_HEADERS}
        lemma = row.get("lemma")
        level = row.get("level")
        unit_no_raw = row.get("unit_no")
        lesson_no_raw = row.get("lesson_no")
        first_page_raw = row.get("first_page")

        if not lemma or not level or unit_no_raw is None or lesson_no_raw is None or first_page_raw is None:
            _stage_unmapped(
                session,
                lemma=str(lemma) if lemma else None,
                level=str(level) if level else None,
                lesson=None,
                first_page=int(first_page_raw) if first_page_raw else None,
                reason="missing_required_fields",
                metadata=_metadata_from_row(row),
            )
            stats["unmapped"] += 1
            continue

        lemma = str(lemma).strip()
        level = str(level).strip()
        lesson = lesson_code(unit_no_raw, lesson_no_raw)

        if level not in lessons_by_level:
            lessons_by_level[level] = _page_lesson_map(session, level)

        lesson_rows = lessons_by_level[level]
        lesson_obj = next((ls for ls in lesson_rows if ls.lesson == lesson), None)
        if not lesson_obj:
            _stage_unmapped(
                session,
                lemma=lemma,
                level=level,
                lesson=lesson,
                first_page=int(first_page_raw),
                reason="unknown_lesson",
                metadata=_metadata_from_row(row),
            )
            stats["unmapped"] += 1
            continue

        first_page = int(first_page_raw)
        if not (lesson_obj.page_start <= first_page <= lesson_obj.page_end):
            _stage_unmapped(
                session,
                lemma=lemma,
                level=level,
                lesson=lesson,
                first_page=first_page,
                reason="first_page_out_of_lesson_range",
                metadata=_metadata_from_row(row),
            )
            stats["unmapped"] += 1
            continue

        norm = normalize_key(lemma)
        if not norm:
            stats["unmapped"] += 1
            continue

        pl = lesson_for_page(lesson_rows, first_page)
        valid_occurrences: list[tuple[Lesson, int]] = [(pl, first_page)] if pl else []

        if not valid_occurrences:
            _stage_unmapped(
                session,
                lemma=lemma,
                level=level,
                lesson=lesson,
                first_page=first_page,
                reason="no_valid_page_mapping",
                metadata=_metadata_from_row(row),
            )
            stats["unmapped"] += 1
            continue

        valid_occurrences.sort(key=lambda x: x[0].order_index * 10000 + x[1])

        meta_json = json.dumps(_metadata_from_row(row), ensure_ascii=False) or None
        source_type = str(row.get("source_type") or "glossary_index_ocr")

        lex = session.query(Lexeme).filter(Lexeme.normalized_lemma == norm).one_or_none()
        if not lex:
            earliest_lesson, earliest_page = valid_occurrences[0]
            lex = Lexeme(
                lemma=lemma,
                normalized_lemma=norm,
                gloss_en=str(row.get("gloss_en") or "") or None,
                item_type="expression" if " " in lemma or len(norm) > 8 else "vocab",
                source_type=source_type,
                review_status="approved",
                first_level=level,
                first_lesson=earliest_lesson.lesson,
                first_page=earliest_page,
                first_order_index=earliest_lesson.order_index,
                raw_source_metadata=meta_json,
            )
            session.add(lex)
            session.flush()
            stats["imported_lexemes"] += 1
        else:
            if lex.gloss_en is None and row.get("gloss_en"):
                lex.gloss_en = str(row.get("gloss_en"))

        for pl, page in valid_occurrences:
            exists = (
                session.query(Occurrence)
                .filter(
                    Occurrence.lexeme_id == lex.id,
                    Occurrence.level == level,
                    Occurrence.lesson == pl.lesson,
                    Occurrence.page == page,
                )
                .one_or_none()
            )
            if not exists:
                session.add(
                    Occurrence(
                        lexeme_id=lex.id,
                        level=level,
                        lesson=pl.lesson,
                        page=page,
                        order_index=pl.order_index,
                        source_type=source_type,
                        review_status="approved",
                        raw_source_metadata=meta_json,
                    )
                )
                stats["occurrences"] += 1

        all_occs = (
            session.query(Occurrence)
            .filter(Occurrence.lexeme_id == lex.id)
            .order_by(Occurrence.order_index, Occurrence.page)
            .all()
        )
        if all_occs:
            first = all_occs[0]
            lex.first_level = first.level
            lex.first_lesson = first.lesson
            lex.first_page = first.page
            lex.first_order_index = first.order_index

        _ensure_surface_form(session, lex.id, lemma, norm, "observed")

        # 다단어 표현: 조사 제거 원형 시퀀스 key도 surface form으로 저장
        # 예: "수업이 끝나다" -> "수업끝나다", "값이 비싸다" -> "값비싸다"
        lemma_key = compute_lemma_key(lemma)
        if lemma_key and lemma_key != norm:
            _ensure_surface_form(session, lex.id, lemma, lemma_key, "generated")
        session.flush()

    session.commit()
    wb.close()
    return stats


def _ensure_surface_form(
    session: Session, lexeme_id: int, surface: str, normalized: str, source: str
) -> None:
    exists = (
        session.query(SurfaceForm)
        .filter(
            SurfaceForm.lexeme_id == lexeme_id,
            SurfaceForm.normalized_surface == normalized,
        )
        .one_or_none()
    )
    if not exists:
        session.add(
            SurfaceForm(
                lexeme_id=lexeme_id,
                surface=surface,
                normalized_surface=normalized,
                source=source,
            )
        )


def compute_lemma_key(lemma: str) -> str | None:
    """표제어 → 내용어 원형 시퀀스 key (import 시 surface form 생성용)."""
    from kvocab_core.morph import KoreanMorphAnalyzer

    morph = KoreanMorphAnalyzer()
    if morph.backend_name != "kiwi":
        return None
    segs = build_match_segments(morph.analyze(lemma))
    if not segs:
        return None
    return segment_key(segs) or None


def main(argv: list[str] | None = None) -> None:
    import sys

    from kvocab_core.database import init_db

    path = Path(argv[1] if argv and len(argv) > 1 else sys.argv[1])
    factory = init_db()
    with factory() as session:
        from kvocab_core.seed import seed_levels_and_lessons

        seed_levels_and_lessons(session, "2A")
        stats = import_vocabulary_xlsx(session, path)
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
