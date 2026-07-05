"""Merge per-level xlsx files under vocabulary/ into data/seed/vocabulary_database.xlsx."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from kvocab_core.config import PACKAGE_ROOT, VOCABULARY_DIR, compute_order_index

VOCAB_KEEP = [
    "level",
    "unit_no",
    "lesson_no",
    "unit_topic",
    "unit_title",
    "first_page",
    "order_index",
    "lemma",
    "gloss_en",
    "source_type",
]
LEVEL_MAP_KEEP = [
    "level",
    "unit_no",
    "lesson_no",
    "unit_topic",
    "unit_title",
    "page_start",
    "page_end",
    "order_index",
]


def merge_vocabulary_xlsx(
    sources: list[Path] | None = None,
    output: Path | None = None,
) -> Path:
    src_dir = VOCABULARY_DIR
    files = sources or sorted(src_dir.glob("snu*_level_mapped_vocabulary.xlsx"))
    if not files:
        raise FileNotFoundError(f"No level xlsx files in {src_dir}")

    all_vocab: list[list] = []
    all_level: list[list] = []

    for fp in files:
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        for sn, bucket in [("Vocabulary", all_vocab), ("Level_Map", all_level)]:
            ws = wb[sn]
            header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col = {name: i for i, name in enumerate(header)}
            for row_vals in ws.iter_rows(min_row=2, values_only=True):
                if not row_vals or not row_vals[col["level"]]:
                    continue
                row = {h: row_vals[col[h]] for h in header}
                level = str(row["level"]).strip()
                unit_no, lesson_no = int(row["unit_no"]), int(row["lesson_no"])
                oi = compute_order_index(level, unit_no, lesson_no)
                if sn == "Vocabulary":
                    bucket.append(
                        [
                            level,
                            unit_no,
                            lesson_no,
                            row["unit_topic"],
                            row["unit_title"],
                            row["first_page"],
                            oi,
                            row["lemma"],
                            row["gloss_en"],
                            row["source_type"],
                        ]
                    )
                else:
                    bucket.append(
                        [
                            level,
                            unit_no,
                            lesson_no,
                            row["unit_topic"],
                            row["unit_title"],
                            row["page_start"],
                            row["page_end"],
                            oi,
                        ]
                    )
        wb.close()

    out = output or PACKAGE_ROOT / "data" / "seed" / "vocabulary_database.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    if out.exists():
        out.unlink()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Vocabulary")
    ws.append(VOCAB_KEEP)
    for row in all_vocab:
        ws.append(row)
    ws = wb.create_sheet("Level_Map")
    ws.append(LEVEL_MAP_KEEP)
    for row in all_level:
        ws.append(row)
    wb.save(out)
    return out


def main() -> None:
    out = merge_vocabulary_xlsx()
    print(f"merged -> {out}")


if __name__ == "__main__":
    main()
